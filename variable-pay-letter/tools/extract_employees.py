#!/usr/bin/env python3
"""Extract the fields the variable pay letter needs from the HR master database.

The master workbook carries PAN, Aadhaar, bank accounts, home addresses, personal
email and phone numbers. None of that belongs in a letter, and none of it belongs
in this repository, so this script copies out only the seven fields the letter
prints and drops everything else on the floor.

    python tools/extract_employees.py path/to/Master_Database.xlsx
"""

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Sheet names HR has used so far. The header sniff below is the real safety net:
# the sheet was renamed once already, and hardcoding a name breaks every time.
KNOWN_SHEETS = ("Active EM", "Active Employees List")

# Header label -> key in the emitted record. Matching is on the normalised
# header text so stray spacing or capitalisation in the workbook is tolerated.
COLUMNS = {
    "emp code": "code",
    "name": "name",
    "designation": "designation",
    "date of joining": "doj",
    "department": "department",
    "sub-department": "subDepartment",
    "location": "location",
}

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "employees.json"


def normalise(value):
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def clean(value):
    """Collapse whitespace; the workbook has plenty of trailing spaces."""
    return re.sub(r"\s+", " ", str(value)).strip()


def as_iso_date(value):
    """Dates arrive as datetimes, but hand-typed rows can be strings."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def find_sheet(workbook):
    """The employee sheet, by name if we recognise it, otherwise by its headers."""
    for name in KNOWN_SHEETS:
        if name in workbook.sheetnames:
            return workbook[name]

    for sheet in workbook.worksheets:
        labels = {normalise(sheet.cell(1, column).value)
                  for column in range(1, min(sheet.max_column, 60) + 1)}
        if "emp code" in labels:
            print(f"  note: using sheet {sheet.title!r}, matched on its headers")
            return sheet

    sys.exit("No employee sheet found: none named "
             f"{list(KNOWN_SHEETS)} and none with an 'Emp Code' header. "
             f"Sheets present: {workbook.sheetnames}")


def main(source):
    workbook = load_workbook(source, data_only=True, read_only=False)
    sheet = find_sheet(workbook)

    headers = {}
    for column in range(1, sheet.max_column + 1):
        key = COLUMNS.get(normalise(sheet.cell(1, column).value))
        # First match wins: "Location" must not be overwritten by "Sub-Location".
        if key and key not in headers:
            headers[key] = column

    missing = set(COLUMNS.values()) - set(headers)
    if missing:
        sys.exit(f"Missing expected columns: {sorted(missing)}")

    employees = []
    seen = set()
    for row in range(2, sheet.max_row + 1):
        record = {}
        for key, column in headers.items():
            value = sheet.cell(row, column).value
            if value is None:
                record[key] = ""
            elif key == "doj":
                record[key] = as_iso_date(value)
            else:
                record[key] = clean(value)

        code = record["code"].upper()
        if not code or not record["name"]:
            continue
        if code in seen:
            print(f"  warning: duplicate employee code {code} on row {row}, keeping the first")
            continue
        seen.add(code)
        record["code"] = code
        employees.append(record)

    employees.sort(key=lambda employee: employee["code"])
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(employees, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")

    print(f"Wrote {len(employees)} employees to {OUT.relative_to(ROOT)}")
    for key in ("designation", "doj", "department", "subDepartment", "location"):
        blank = sum(1 for employee in employees if not employee[key])
        if blank:
            print(f"  note: {blank} record(s) have no {key}; the letter will omit that line")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    main(sys.argv[1])
