"""Turn the Sales roster workbooks into the tree the page renders.

Reporting lines come from the Reporting Manager column, which holds names
rather than employee codes, so managers are matched by normalised name with a
token-subset fallback for spelling variants. People are keyed by employee code
throughout, so colleagues who share a name stay distinct.
"""
import openpyxl, collections, re, json

EMP = 'roster.xlsx'              # current roster: adds Sub-Location for everyone
GRADES_FROM = '/root/.claude/uploads/e357e2b7-101a-5852-81e0-5d4ad216b53a/33237778-Sales_Employees_1.xlsx'
INT = '/root/.claude/uploads/e357e2b7-101a-5852-81e0-5d4ad216b53a/c5856b3b-Sales_Interns.xlsx'
TERR = 'territory.xlsx'          # city -> zone -> RM coverage map
MASTER = 'masterfile.xlsx'       # validated field-force file: channel, city, outlet
DEALERS = 'dealers.xlsx'         # dealer master: dealership -> ASM -> RM -> zone

# There is no channel concept in Inside Sales, Sales Ops, Expansion or Brands,
# so those lines never show a channel figure. Everyone else may: a Regional
# Manager recorded under Sales Central still runs a retail region.
NO_CHANNEL = {'Inside Sales', 'Sales Ops', 'Expansion', 'Network Expansion',
              'Brands', 'Sales and Marketing'}

# Clear data errors, corrected here rather than left to mislead. Each is a
# single place recorded two ways, or a value that is not a place at all.
CITY_FIX = {
    'kolkata(east)': 'Kolkata',        # same city as Kolkata
    'bengaluru': 'Bangalore',          # same city, both spellings in use
    'new agra': 'Agra',                # same city
    'new delhi': 'Delhi',              # same city as the territory map's Delhi
    'manager strategic sourcing and procurement': '',   # a job title in a city field
    'coco': '',                        # an outlet model, not a city
}

# lines that report outside Sales leadership and are out of scope for this chart
EXCLUDE_LINES = ['Raman Pandey', 'Vijay Malik', 'Jashanjot Singh']

# the roster splits retail by vehicle class; the chart treats it as one function
RETAIL = {'3W Retail Sales', '4W Retail Sales', 'Retail Sales'}

# The South seat, open since Amitabh Singh resigned, has been filled.
SOUTH_RM = 'Rohit Madhavan'
# people the roster still records against the departed RM move to the new one
LEGACY_SOUTH_RM = 'amitabhsingh'
# and this Senior ASM, parked under the AVP while the seat was open, moves back
MOVED_TO_SOUTH_RM = ['Ravi Prakash']

# The roster records these people by grade (General Manager and such); the org
# role they actually hold is Regional Manager. Confirmed against the RM list.
REGIONAL_MANAGERS = ['Vinay Binu', 'Rachit Sharma', 'Vikas Singh Bisht',
                     'Arup Roy Chowdhury', 'Subesh Mukherjee', 'Siddhartha Sharma',
                     'Gaurav Bhardwaj', SOUTH_RM]

# what each function head actually owns — the roster designation alone
# ("General Manager") does not say which part of Sales that is
PORTFOLIOS = {
    'Ratanmani Mohit': 'Retail Sales & Brands',
    'Shashank': 'Sales Operations',
    'Gaurav Bhardwaj': 'Regional Management · Sales Central · Special Projects',
    'Ritesh Roy': 'Sales Expansion',
}

# the territory and dealer sheets name RMs by first name only
TERRITORY_OWNER = {
    'vinay': 'Vinay Binu', 'rachit': 'Rachit Sharma', 'vikas': 'Vikas Singh Bisht',
    'arup': 'Arup Roy Chowdhury', 'subesh': 'Subesh Mukherjee',
    'siddhartha': 'Siddhartha Sharma',
    # territory sits with the Regional Manager. Amit Vishwakarma is an AGM under
    # Vikas Singh Bisht, so Faridabad and Sonipat belong to Vikas.
    'amit': 'Vikas Singh Bisht',
    # the sheets listed the South cities under Mohit while he covered the seat,
    # and older rows still name the departed RM
    'mohit': SOUTH_RM,
    'amitabh': SOUTH_RM,
}

def fix_city(c):
    return CITY_FIX.get(clean(c).lower(), clean(c))

# these two sit across the whole department, so a single state would mislead
HIDE_LOCATION = ['Vani Rikhy Mehra', 'Ratanmani Mohit']

# field roles counted per Regional Manager
IS_TSM = lambda t: 'Territory Sales Manager' in t          # includes Senior TSM
IS_TSE = lambda t: 'Territory Sales Executive' in t

norm = lambda n: re.sub(r'[^a-z]', '', (n or '').lower())
clean = lambda s: (str(s).strip() if s is not None else '')

def merge_sub(s):
    s = clean(s)
    return 'Retail Sales' if s in RETAIL else s


# ── read ──────────────────────────────────────────────────────────────────
# the current roster dropped the Grade column; carry it over from the previous
# extract where the employee code matches
GRADE = {}
for r in openpyxl.load_workbook(GRADES_FROM, data_only=True).active.iter_rows(min_row=2, values_only=True):
    if r[0] and clean(r[6]):
        GRADE[clean(r[0])] = clean(r[6])

emps = []
for r in openpyxl.load_workbook(EMP, data_only=True).active.iter_rows(min_row=2, values_only=True):
    if not (r[2] and clean(r[2]) == 'Sales'):
        continue
    code = clean(r[0])
    emps.append({'id': code, 'name': clean(r[1]), 'sub': merge_sub(r[3]),
                 'title': clean(r[4]), 'band': clean(r[5]),
                 'grade': GRADE.get(code, ''), 'doj': '',
                 'mgr': clean(r[6]), 'loc': clean(r[7]),
                 'city': fix_city(r[8]), 'intern': False})

interns = []
for r in openpyxl.load_workbook(INT, data_only=True).active.iter_rows(min_row=2, values_only=True):
    if not r[0]:
        continue
    interns.append({'id': clean(r[0]), 'name': clean(r[1]), 'sub': merge_sub(r[3]),
                    'title': 'Intern', 'band': 'Intern', 'grade': '', 'doj': '',
                    'mgr': clean(r[5]), 'loc': '', 'city': '', 'intern': True})

people = emps + interns
by_name = {}
for p in emps:
    by_name.setdefault(norm(p['name']), p['id'])

ALIAS = {'naitiksrivastav': 'naitiksrivastava', 'arupchoudhary': 'aruproychowdhury'}
OUTSIDE = {'sauravkumar', 'ashishtandon', 'abhishekmalik'}

def resolve(m):
    k = ALIAS.get(norm(m), norm(m))
    if k == LEGACY_SOUTH_RM:         # departed South RM — his line moves to the new one
        return by_name.get(norm(SOUTH_RM))
    if k in by_name:
        return by_name[k]
    if k in OUTSIDE:
        return None
    toks = set(re.findall(r'[a-z]+', (m or '').lower()))
    c = [p for p in emps if toks and toks <= set(re.findall(r'[a-z]+', p['name'].lower()))]
    return c[0]['id'] if len(c) == 1 else None

for p in people:
    p['mgrId'] = resolve(p['mgr']) if p['mgr'] else None

south = by_name.get(norm(SOUTH_RM))
for p in people:
    if p['name'] in MOVED_TO_SOUTH_RM and south:
        p['mgrId'] = south

byid = {p['id']: p for p in people}
kids = collections.defaultdict(list)
for p in people:
    if p['mgrId'] and p['mgrId'] != p['id']:
        kids[p['mgrId']].append(p['id'])

# ── drop the out-of-scope lines, whole subtree each ───────────────────────
def subtree(i, acc=None):
    acc = acc if acc is not None else set()
    if i in acc:
        return acc
    acc.add(i)
    for c in kids[i]:
        subtree(c, acc)
    return acc

dropped = set()
for nm in EXCLUDE_LINES:
    if norm(nm) in by_name:
        dropped |= subtree(by_name[norm(nm)])

people = [p for p in people if p['id'] not in dropped]
byid = {p['id']: p for p in people}
for k in list(kids):
    if k in dropped:
        del kids[k]
    else:
        kids[k] = [c for c in kids[k] if c not in dropped]

# ── the South seat, now filled ────────────────────────────────────────────
if south:
    byid[south]['note'] = ('Joined to lead the South region. The seat was open '
                           'after Amitabh Singh’s resignation, covered in the '
                           'interim by Ratanmani Mohit.')

RM_IDS = {by_name[norm(n)] for n in REGIONAL_MANAGERS if norm(n) in by_name}
for i in RM_IDS:
    byid[i]['rm'] = True
    byid[i]['roster_title'] = byid[i]['title']
    byid[i]['title'] = 'Regional Manager'

# ── field-force masterfile ────────────────────────────────────────────────
# Adds what the roster does not carry: whether someone sells through the
# channel or in the field, the city they work, and the kind of outlet.
MASTER_ROWS = {}
for r in openpyxl.load_workbook(MASTER, data_only=True).active.iter_rows(min_row=2, values_only=True):
    code = clean(r[0])
    if not code:
        continue
    cat = clean(r[13]).replace('Shawroom', 'Showroom')       # sheet typo
    MASTER_ROWS[code] = {'chan': clean(r[6]), 'city': clean(r[15]),
                         'cat': cat, 'flagged': clean(r[16]) == 'Incorrect Mapping'}

# ── dealer master ─────────────────────────────────────────────────────────
# A dealership belongs to an ASM, who belongs to an RM. Offboarded dealers are
# counted separately so a live network figure stays live.
DEALERS_BY = collections.defaultdict(lambda: {'live': 0, 'off': 0})
DEALER_LIST = collections.defaultdict(list)   # owner id -> [{name, city, model, live}]
DEALER_META = {'model': collections.Counter(), 'status': collections.Counter(),
               'tier': collections.Counter()}
_dseen = set()
for r in openpyxl.load_workbook(DEALERS, data_only=True)['Sheet1'].iter_rows(min_row=6, values_only=True):
    name = clean(r[12])
    if not name or name in _dseen:
        continue
    _dseen.add(name)
    status, model, tier = clean(r[10]), clean(r[9]), clean(r[8])
    DEALER_META['model'][model] += 1
    DEALER_META['status'][status] += 1
    DEALER_META['tier'][tier] += 1
    live = status != 'Offboarded'
    for who, col in (('rm', 2), ('asm', 3)):
        val = clean(r[col])
        if not val or val == '-':
            continue
        target = TERRITORY_OWNER.get(val.lower(), val) if who == 'rm' else val
        oid = by_name.get(norm(target))
        if oid:
            DEALERS_BY[oid]['live' if live else 'off'] += 1
            DEALER_LIST[oid].append({'name': name, 'city': fix_city(r[7]) or clean(r[6]),
                                     'model': model, 'live': live})

# ── territory map ─────────────────────────────────────────────────────────
TERRITORY = collections.defaultdict(list)     # owner id -> [{city, zone}]
_seen = set()
for r in openpyxl.load_workbook(TERR, data_only=True).active.iter_rows(min_row=2, values_only=True):
    city, zone, owner = clean(r[0]), clean(r[1]), clean(r[2])
    if not city or not owner:
        continue
    city = fix_city(city)
    if not city:
        continue
    key = (city.lower(), zone.lower(), owner.lower())
    if key in _seen:                          # the sheet repeats a few rows verbatim
        continue
    _seen.add(key)
    target = TERRITORY_OWNER.get(owner.lower())
    if not target:
        continue
    oid = by_name.get(norm(target))
    if oid:
        TERRITORY[oid].append({'city': city, 'zone': zone})

HIDE_LOC_IDS = {by_name[norm(n)] for n in HIDE_LOCATION if norm(n) in by_name}
PORTFOLIO_BY_ID = {by_name[norm(n)]: v for n, v in PORTFOLIOS.items() if norm(n) in by_name}

ROOT = by_name[norm('Vani Rikhy Mehra')]

# ── roll-ups ──────────────────────────────────────────────────────────────
def rollup(i, seen=None):
    """(employees, interns) in this line, counting the person named."""
    seen = seen if seen is not None else set()
    if i in seen:
        return 0, 0
    seen.add(i)
    p = byid[i]
    emp = 0 if (p.get('vacant') or p['intern']) else 1
    itn = 1 if p['intern'] else 0
    for c in kids[i]:
        ce, ci = rollup(c, seen)
        emp += ce
        itn += ci
    return emp, itn

def field_roles(i, acc=None):
    """Territory managers and executives anywhere in this line."""
    acc = acc if acc is not None else {'tsm': 0, 'tse': 0}
    t = byid[i]['title']
    if IS_TSM(t): acc['tsm'] += 1
    elif IS_TSE(t): acc['tse'] += 1
    for c in kids[i]:
        field_roles(c, acc)
    return acc

def channel_split(i, acc=None):
    acc = acc if acc is not None else {'field': 0, 'chan': 0}
    m = MASTER_ROWS.get(i)
    if m:
        acc['chan' if m['chan'] == 'Channel Sales' else 'field'] += 1
    for c in kids[i]:
        channel_split(c, acc)
    return acc

def locations(i, acc=None):
    acc = acc if acc is not None else collections.Counter()
    if byid[i].get('loc'):
        acc[byid[i]['loc']] += 1
    for c in kids[i]:
        locations(c, acc)
    return acc

ABBR = [('Assistant Vice President', 'AVP'), ('Vice President', 'VP'),
        ('Senior Territory Sales Manager', 'Sr TSM'), ('Territory Sales Manager', 'TSM'),
        ('Territory Sales Executive', 'TSE'), ('Senior Area Sales Manager', 'Sr ASM'),
        ('Area Sales Manager', 'ASM'), ('Key Accounts Manager', 'KAM'),
        ('Key Account Manager', 'KAM'), ('Assistant General Manager', 'AGM'),
        ('General Manager', 'GM'), ('Regional Manager', 'RM'),
        ('Senior Sales Executive', 'Sr Sales Exec'), ('Sales Executive', 'Sales Exec'),
        ('Senior Executive- Accounts', 'Sr Exec (Accts)'), ('Senior Executive', 'Sr Exec'),
        ('Senior Manager- Inside Sales', 'Sr Mgr'), ('Senior Manager', 'Sr Mgr'),
        ('Assistant Manager - New Product Sales', 'Asst Mgr'),
        ('Assistant Manager', 'Asst Mgr'), ('Senior Pogram Manager', 'Sr Program Mgr'),
        ('Program Manager', 'Program Mgr'), ('Data Analyst Executive', 'Data Analyst'),
        ('State Head-Sales', 'State Head'), ('Regional Trainer', 'Trainer'),
        ('Manager-Network Expansion', 'Mgr'), ('Manager- BTL Marketing', 'Mgr'),
        ('Intern', 'Intern')]

def abbr(t):
    t = clean(t)
    for long, short in ABBR:
        if t.startswith(long):
            return short
    return t


def node(i):
    """A manager becomes a tree node; their individual contributors become that
    node's team, so the tree stays a tree instead of a 36-wide row of leaves."""
    p = byid[i]
    emp, itn = rollup(i)
    mgr_kids = [c for c in kids[i] if kids[c]]
    ics = [c for c in kids[i] if not kids[c]]

    d = {'id': i, 'name': p['name'], 'title': p['title'], 'short': abbr(p['title']),
         'sub': p['sub'], 'grade': p['grade'], 'doj': p['doj'], 'loc': p.get('loc', ''),
         'emp': emp, 'interns': itn, 'direct': len(kids[i])}
    if p.get('vacant'):
        d['vacant'] = True
        d['note'] = p['note']
    if p.get('note'):
        d['note'] = p['note']
    if i in PORTFOLIO_BY_ID:
        d['portfolio'] = PORTFOLIO_BY_ID[i]
    cs = channel_split(i)
    if (cs['chan'] or cs['field']) and p['sub'] not in NO_CHANNEL:
        d['field'], d['chan'] = cs['field'], cs['chan']
    if i in DEALERS_BY:
        d['dealers'] = DEALERS_BY[i]['live']
        if DEALERS_BY[i]['off']:
            d['dealersOff'] = DEALERS_BY[i]['off']
        d['dealerList'] = sorted(DEALER_LIST[i],
                                 key=lambda x: (not x['live'], x['city'], x['name']))
    if i in TERRITORY:
        t = sorted(TERRITORY[i], key=lambda x: x['city'])
        z = collections.Counter(x['zone'] for x in t)
        d['cities'] = [x['city'] for x in t]
        d['zones'] = [{'name': k, 'n': v} for k, v in z.most_common()]

    if p.get('rm'):
        d['rm'] = True
        d['rosterTitle'] = p.get('roster_title', '')
    f = field_roles(i)
    if f['tsm'] or f['tse']:
        d['tsm'], d['tse'] = f['tsm'], f['tse']

    if i in HIDE_LOC_IDS:
        d['loc'] = ''
    else:
        top = locations(i).most_common(3)
        if top:
            d['locs'] = [{'name': k, 'n': v} for k, v in top]

    if ics:
        mem = sorted((byid[c] for c in ics), key=lambda x: (x['intern'], x['name']))
        d['team'] = {
            'emp': sum(0 if m['intern'] else 1 for m in mem),
            'interns': sum(1 for m in mem if m['intern']),
            'titles': collections.Counter(abbr(m['title']) for m in mem).most_common(),
            'members': [dict({'name': m['name'], 'title': m['title'],
                              'loc': m.get('city') or m.get('loc', ''),
                              'intern': m['intern'], 'grade': m['grade']},
                             **({'chan': 'C'} if MASTER_ROWS.get(m['id'], {}).get('chan')
                                == 'Channel Sales' else {}))
                        for m in mem]}

    if mgr_kids:
        ch = [node(c) for c in mgr_kids]
        ch.sort(key=lambda x: -(x['emp'] + x['interns']))
        d['children'] = ch
    return d


tree = node(ROOT)

emp_total = sum(1 for p in people if not p['intern'] and not p.get('vacant'))
int_total = sum(1 for p in people if p['intern'])

meta = {
    'employees': emp_total, 'interns': int_total, 'total': emp_total + int_total,
    'subDepts': collections.Counter(p['sub'] for p in people
                                    if not p['intern'] and not p.get('vacant')).most_common(),
    'grades': sorted(collections.Counter(p['grade'] for p in people if p['grade']).items()),
    'locations': collections.Counter(p['loc'] for p in people if p.get('loc')).most_common(),
    'topCities': collections.Counter(p['city'] for p in people if p.get('city')).most_common(12),
    'cities': sum(len(v) for v in TERRITORY.values()),
    'channel': [('Field sales', sum(1 for v in MASTER_ROWS.values()
                                    if v['chan'] == 'Field Sales')),
                ('Channel sales', sum(1 for v in MASTER_ROWS.values()
                                      if v['chan'] == 'Channel Sales'))],
    'outlets': collections.Counter(v['cat'] for v in MASTER_ROWS.values()
                                   if v['cat']).most_common(),
    'dealerModel': DEALER_META['model'].most_common(),
    'dealerStatus': DEALER_META['status'].most_common(),
    'dealerTier': DEALER_META['tier'].most_common(),
    'dealersLive': sum(1 for k, v in DEALER_META['status'].items()
                       if k != 'Offboarded' for _ in range(v)),
    'dealersAll': sum(DEALER_META['status'].values()),
    'zones': collections.Counter(x['zone'] for v in TERRITORY.values()
                                 for x in v).most_common(),
    'excluded': EXCLUDE_LINES,
}
json.dump({'tree': tree, 'meta': meta}, open('org.json', 'w'), indent=1)

# ── checks ────────────────────────────────────────────────────────────────
def placed(n):
    t = 0 if n.get('vacant') else 1
    t += len(n.get('team', {}).get('members', []))
    return t + sum(placed(c) for c in n.get('children', []))

def mgr_nodes(n):
    return 1 + sum(mgr_nodes(c) for c in n.get('children', []))

def depth(n):
    return 1 + max([depth(c) for c in n.get('children', [])] or [0])

print(f'employees {emp_total} + interns {int_total} = {emp_total + int_total}')
print('people placed in tree :', placed(tree))
print('dropped with excluded :', len(dropped))
print('manager nodes / depth :', mgr_nodes(tree), '/', depth(tree))
print('root line             :', tree['emp'], 'employees /', tree['interns'], 'interns')
print('regional managers     :', len(RM_IDS))
def rmline(n, out=None):
    out = out if out is not None else []
    if n.get('rm'):
        out.append((n['name'], n.get('tsm', 0), n.get('tse', 0), n['emp']))
    for c in n.get('children', []): rmline(c, out)
    return out
for nm, tsm, tse, emp in rmline(tree):
    print(f'   {nm:32s} {emp:4d} in line | {tsm:3d} TSM | {tse:3d} TSE')
print('cities mapped         :', sum(len(v) for v in TERRITORY.values()),
      'across', len(TERRITORY), 'owners')
print('dealerships           :', sum(DEALER_META['status'].values()),
      '| live:', sum(v for k, v in DEALER_META['status'].items() if k != 'Offboarded'))
print('masterfile rows       :', len(MASTER_ROWS),
      '| joined to roster:', sum(1 for c in MASTER_ROWS if c in byid))
print('channel sellers       :', sum(1 for v in MASTER_ROWS.values()
                                     if v['chan'] == 'Channel Sales'))
print('sub-departments       :', meta['subDepts'])
